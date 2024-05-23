import os
from openpyxl import load_workbook

def find_xlsx_files(path):
    # 搜索指定目录下的所有xlsx文件
    xlsx_files = [os.path.join(path, f) for f in os.listdir(path) if f.endswith('.xlsx')]
    return xlsx_files

def get_sheets_from_xlsx(file_path):
    # 加载Excel文件
    workbook = load_workbook(filename=file_path, data_only=True)
    # 获取所有sheet的名字
    sheets = workbook.sheetnames
    return sheets

def main():
    # 指定目录
    path = "/home/liuxinn/RLCLAL/self/learn_from_md/deir-main/src/logs/Results"

    # 获取目录下的所有Excel文件
    xlsx_files = find_xlsx_files(path)

    # 输出每个文件的sheets
    for file_path in xlsx_files:
        sheets = get_sheets_from_xlsx(file_path)
        print(f"File: {file_path} has the following sheets: {sheets}")




if __name__ == "__main__":
    main()
